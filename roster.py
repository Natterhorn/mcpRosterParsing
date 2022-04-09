from pickle import FALSE, TRUE
from openpyxl import load_workbook
from rapidfuzz import fuzz
from rapidfuzz import process
from rapidfuzz.fuzz import ratio, partial_ratio, token_sort_ratio, partial_token_sort_ratio, token_set_ratio
import re
import nicknameRef as nRef

def findMatch(strVal, printResults):
    if strVal.lower() in nRef.nicknames:
        strVal = nRef.nicknames[strVal.lower()]

    #Remove excel newlines
    compareStr = strVal.replace("_x000D_", "")
    #Remove hyphen+whitespace at start of string
    compareStr = re.sub("- (?=\w)", "", compareStr)
    #Remove brackets with characters between them (ex [B], [18])
    compareStr = re.sub("\[(?=\w)\]", "", compareStr)
    #Remove whitespace from front and back
    compareStr = compareStr.strip()

    #Use most specific scorer first, then try more and more permissive ones
    fuzzyMatch = process.extractOne(compareStr, joinedlist, scorer=ratio, score_cutoff=100)
    sc = 'ratio'

    if fuzzyMatch == None:
        fuzzyMatch = process.extractOne(compareStr, joinedlist, scorer=token_sort_ratio, score_cutoff=85)
        sc = 'token_sort'

    if fuzzyMatch == None:
        fuzzyMatch = process.extractOne(compareStr, joinedlist, scorer=token_set_ratio, score_cutoff=85)
        sc = 'token_set_ratio'

    if fuzzyMatch == None:
        fuzzyMatch = process.extractOne(compareStr, joinedlist, scorer=partial_ratio, score_cutoff=85)
        sc = 'partial'

    #For testing, print results of the matching
    if printResults == TRUE:
        if fuzzyMatch != None:
            print(compareStr + ":" + fuzzyMatch[0] + "(" + str(fuzzyMatch[1]) + ")|" + sc)
        else:
            print(compareStr + "|")

    return fuzzyMatch

# Give the location of the file 
path = "event_reg.xlsx"

# Join the list of all card names from helper file
joinedlist = nRef.characters.union(nRef.tactics.union(nRef.extracts.union(nRef.secures)))

#fMatch = findMatch("Cpt America Steve", TRUE)
#exit()

wb = load_workbook(path)
sheet_obj = wb.active 

for i in range(2, sheet_obj.max_row+1):
#for i in range(5, 6):
    sheetName = "Roster " + str(i)
    # Create a sheet for each roster row
    if sheetName not in wb.sheetnames:
        wb.create_sheet(sheetName)
    roster_sheet = wb[sheetName]

    cell_obj1 = sheet_obj.cell(row = i, column = 3)  
    j = 1

    # Break up cell contents by newlines, throw each through the fuzzymatcher
    for lineStr in cell_obj1.value.splitlines():
        cell_obj2 = roster_sheet.cell(row = j, column = 1)

        match = findMatch(lineStr, FALSE)
        
        #If a match exists, use that. Otherwise, use unmodified
        if match != None:
            cell_obj2.value = match[0]
        else:
            cell_obj2.value = lineStr

        #Copy the exact value used to a different column to crossreference output
        cell_obj2 = roster_sheet.cell(row = j, column = 11)
        cell_obj2.value = lineStr   
        j += 1

wb.save(path)